import openpyxl


class ReadDataSheet:
    @staticmethod
    def get_max_rcount(file, sheetname):
        wb= openpyxl.load_workbook(file)
        sh= wb[sheetname]
        return sh.max_row

    @staticmethod
    def get_max_ccount(file,sheetname):
        wb = openpyxl.load_workbook(file)
        sh= wb[sheetname]
        return sh.max_column

    @staticmethod
    def read_data(file, sheetname, rownum, colnum):
        wb= openpyxl.load_workbook(file)
        sh= wb[sheetname]
        return sh.cell(row=rownum, column=colnum).value
